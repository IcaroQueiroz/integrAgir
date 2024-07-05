import pandas as pd
from itertools import combinations
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font

class AppAthenas():
    def athenas(self):
        try:
            df_origem = self.df[self.df['Contabilização'] == 0]
            df_origem['Data aquisição1'] = df_origem['Data aquisição']
            df_origem['Data aquisição2'] = df_origem['Data aquisição']
            df_origem['Valor original da aquisição2'] = df_origem['Valor original da aquisição']
            print("------------def ogigem -------------")
            print(df_origem)
            # Criar uma nova planilha
            self.wb = Workbook()
            # Adicionar uma nova folha
            self.ws = self.wb.active

            # Cabeçalhos
            cabecalhos = [
                "Código do Fornecedor", "Nome Fornecedor", "CNPJ do Fornecedor", "Nota Fiscal", "Quantidade",
                "Patrimônio", "Data da Abertura", "Vida Útil", "Código da Conta Contábil", "Descrição do Bem",
                "Código do Centro de Custo", "Data da Entrada da Compra", "Data da Aquisição", "Valor Total da NF",
                "Base de Cálculo", "Aliquota", "Crédito de ICMS", "ICMS Adicional", "Valor do Mês", "Valor Alocado",
                "Código da Conta de Despesa da Depreciação", "SALDO DE ABERTURA", "CIAP", "ICMS Frete", "ICMS ST",
                "ICMS DIF", "Código da Conta de Depreciação Acumulada", "Valor Depreciado", "Código da Localização",
                "Parcelas CIAP", "Série", "Vida Útil em Meses", "Observações", "Tipo de Depreciação(0 - Fiscal, 1 - IFRS)",
                "Depreciar no Mês Seguinte(0 - Não, 1 - Sim)", "CPC - Natureza da BC Pis/Cofins", "CPC - Identificação dos Bens/Grupos",
                "CPC - Indicador de Origem do Bem", "CPC - CST Pis", "CPC - CST Cofins", "CPC - Indicador da Utilização do Bens",
                "CPC - Conta Contábil", "CPC - Indicador do Número de Parcelas", "CPC - Número de Parcelas",
                "CPC - Parcela a Excluir da B/C"
            ]

            # Adicionar cabeçalhos à nova planilha
            for col_num, cabecalho in enumerate(cabecalhos, start=1):
                coluna_letra = get_column_letter(col_num)
                self.ws[f"{coluna_letra}1"] = f" {coluna_letra} -> {cabecalho};"
                self.ws.column_dimensions[coluna_letra].width = len(cabecalho) + 4  # Ajuste a largura da coluna conforme necessário
                self.ws[f"{coluna_letra}1"].font = Font(bold=True)

            colunas_para_levar = {
                    "Data aquisição": "G",
                    "Data aquisição1": "M",
                    "Data aquisição2": "L",
                    "Taxa de depreciação": "H",
                    "Código da conta": "I",
                    "Nome do bem": "J",
                    "Valor original da aquisição": "N",
                    "Valor original da aquisição2": "V",
                    }

            # Mapeamento de substituição para a coluna "Código da conta"
            substituicoes_codigo_conta = {
                103: ["1.3.5.03.01.00000001", "1.3.5.03.03.00000001", "6.5.1.09.01.00000001"],
                102: ["1.3.5.03.01.00000002", "1.3.5.03.03.00000002", "6.5.1.09.01.00000001"],
                99: ["1.3.5.03.01.00000003", "1.3.5.03.03.00000003", "6.5.1.09.01.00000001"],
                98: ["1.3.5.03.01.00000004", "1.3.5.03.03.00000004", "6.5.1.09.01.00000001"],
                100: ["1.3.5.03.01.00000005", "1.3.5.03.03.00000005", "6.5.1.09.01.00000001"],
                101: ["1.3.5.03.01.00000006", "1.3.5.03.03.00000006", "6.5.1.09.01.00000001"],
                837: ["1.3.5.03.01.00000007", "1.3.5.03.03.00000007", "6.5.1.09.01.00000001"],
                838: ["1.3.5.03.01.00000008", "1.3.5.03.03.00000008", "6.5.1.09.01.00000001"],
                839: ["1.3.5.03.01.00000010", "1.3.5.03.03.00000010", "6.5.1.09.01.00000001"],
                840: ["1.3.5.03.01.00000011", "1.3.5.03.03.00000009", "6.5.1.09.01.00000001"],
                104: ["1.3.5.03.01.00000012", "1.3.5.03.03.00000010", "6.5.1.09.01.00000001"],
                126: ["1.3.7.01.01.00000001", "1.3.7.01.03.00000001", "6.5.1.09.01.00000001"],
                128: ["1.3.7.01.01.00000002", "1.3.7.01.03.00000002", "6.5.1.09.01.00000001"],
                129: ["1.3.7.01.01.00000003", "1.3.7.01.03.00000003", "6.5.1.09.01.00000001"],
                127: ["1.3.7.01.01.00000004", "1.3.7.01.03.00000004", "6.5.1.09.01.00000001"],
                130: ["1.3.7.01.01.00000005", "1.3.7.01.03.00000005", "6.5.1.09.01.00000001"],
                131: ["1.3.7.01.01.00000006", "1.3.7.01.03.00000006", "6.5.1.09.01.00000001"],
                849: ["1.3.7.01.01.00000007", "1.3.7.01.03.00000007", "6.5.1.09.01.00000001"],
                107: ["1.3.7.01.01.00000008", "1.3.7.01.03.00000008", "6.5.1.09.01.00000001"],
            }
      
            # Adicionar dados das colunas da planilha de origem para as colunas correspondentes na planilha de destino
            for coluna_origem, coluna_destino in colunas_para_levar.items():
                if coluna_origem == "Código da conta":
                    for i, valor in enumerate(df_origem[coluna_origem], start=2):
                        valores_substituidos = substituicoes_codigo_conta.get(valor, [valor, ""])
                        valor_tratado = valores_substituidos[0]  # Usar o primeiro valor da lista como valor tratado
                        self.ws[f"{coluna_destino}{i}"] = valor_tratado
                        # Adicionar o segundo valor da lista (se existir) à outra coluna
                        if valores_substituidos[1]:
                            self.ws[f"AA{i}"] = valores_substituidos[1]
                        if valores_substituidos[2]:
                            self.ws[f"U{i}"] = valores_substituidos[2]  

                elif coluna_origem == "Taxa de depreciação":
                    for i, valor in enumerate(df_origem[coluna_origem], start=2):
                        valor = float(valor.replace(',', '.'))
                        if valor != 0:  # Evita divisão por zero
                            valor_tratado = 100 / valor
                            self.ws[f"{coluna_destino}{i}"] = valor_tratado
                        else:
                            self.ws[f"{coluna_destino}{i}"] = 0  # Pode definir outro valor padrão se necessário
                elif coluna_origem in ["Valor original da aquisição", "Valor original da aquisição2"]:  # Verifica se a coluna de origem é "N" ou "V"
                        for i, valor in enumerate(df_origem[coluna_origem], start=2):
                            # Converter o valor para float e substituir a vírgula por ponto
                            valor_numerico = float(valor.replace('.', '').replace(',', '.')) if valor else 0.0
                            self.ws[f"{coluna_destino}{i}"] = valor_numerico
                else:
                    for i, valor in enumerate(df_origem[coluna_origem], start=2):
                        self.ws[f"{coluna_destino}{i}"] = valor
                        self.ws[f"AB{i}"] = 0

            print("--------------------------------------- self.ws-------------------------------------------------")
            print(self.ws)
            print("--------------------------------------- self.wb-------------------------------------------------")
            print(self.wb)


            QMessageBox.warning(self, 'Info', 'CSV carregado com Sucesso')  
        except KeyError as erro:
            QMessageBox.warning(self, 'KeyError', 'Não foi localizado a coluna: '+ str(erro))
        except ValueError as erro:
            QMessageBox.warning(self, 'ValueError', 'Retire os filtros da planilha.    ')                   
        except Exception as erro:
            QMessageBox.warning(self, 'Error', 'Erro ao tentar carregar o aquivo. - '+ str(erro))    
